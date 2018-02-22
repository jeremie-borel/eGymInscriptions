#!/usr/bin/env python 
# -*- coding: utf-8 -*-

import sys, os, copy
import logging, datetime, re
from collections import defaultdict
from openpyxl import load_workbook

POSTFINANCE_PSWD = None
INSCRIPTION_URL = None
try:
    # tries to load a password file
    from pswd import *
except ImportError:
    pass

# disable request https warning
# http://stackoverflow.com/questions/27981545/suppress-insecurerequestwarning-unverified-https-request-is-being-made-in-pytho#28002687
import requests
import requests.packages.urllib3.exceptions as ulib
requests.packages.urllib3.disable_warnings(ulib.InsecureRequestWarning)


_pat = re.compile( r'^[0-9A-Z]{6}-[0-9A-Z]{6}$' )

from PyFileMaker import FMServer
fm = FMServer( url=INSCRIPTION_URL, debug=False )
fm.setDb( 'Inscription' )
fm.setLayout( 'XmlPaiement' )

FACTURE = u'facture'

def parse_excel( filename ):
    wb = load_workbook(filename=filename, read_only=True)
    ws = wb.active
    all_rows = set()
    tpl = []
    for row in ws.iter_rows( min_row=2, min_col=1, max_col=14 ):
        ukey = row[1].value
        assert( ukey not in all_rows )
        all_rows.add(ukey)

        # not used but ensure the column is a date.
        date_facture = row[7].value.date()
        date_paiement = row[10].value
        num_demande = row[13].value
        montant = row[11].value

        resultat = None
        if montant == 0.0:
            resultat = 'NON_PAYE'
        elif montant == 70.0:
            resultat = 'OK'
        elif montant < 70:
            resultat = 'ERREUR'
        assert( resultat is not None )
        assert( _pat.match(num_demande) )
        data = {
            'numeroDemande': num_demande,
            'modePaiement': FACTURE,
            'datePaiement': date_paiement if date_paiement else '',
            'paiementId' : 'f{}'.format(ukey),
            'resultatPaiement': resultat,
        }
        tpl.append( data )
    return tpl

def download_paiements():
    print "Downloading existing reccords from Inscription..."

    resultset = fm.doFind({
        'modePaiement':FACTURE,
    })
    ret = { r.paiementId:r for r in resultset }
    return ret


def main( options ):
    
    items = parse_excel( options.file )

    current = download_paiements()

    created, edited, up_to_date = 0,0,0
    for data in items:
        key = data['paiementId']
        if key not in current.keys():
            print u"Creating payment for", data['numeroDemande']
            fm.doNew( data )
            created += 1
            continue

        r = current[key]
        flag = False
        for k,value in data.items():
            if getattr( r, k ) != value:
                setattr( r, k, value )
                flag = True
        if flag:
            print u"Editing payment for", data['numeroDemande']
            edited += 1
            fm.doEdit( r )
            continue

        up_to_date += 1

    print "Total number of lines in xl file: ", len(items)
    print "Number of payement row created: ", created
    print "Number of edited payement", edited
    print "Number of payement up to date", up_to_date

if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(
        description=u"""

        """,
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
        
    parser.add_argument( '-f', '--file', dest="file",
                            default=None, required=True,
                            help='Fichier xl de SAP.' )

    args = parser.parse_args()

    main( args )
